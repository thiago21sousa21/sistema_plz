import os
import shutil
import pandas as pd
import re
from openpyxl import load_workbook
from fiscais import FISCAIS
from dados_camera import DADOS_CAMERA
from config import PLZDATA_FILE, DIARIOVIDEO_FILE, AUTO_TEMPLATE_FILE, AUTO_OUTPUT_FILE


try:
    import win32com.client as win32  # Importa pywin32 (somente no Windows)
except ImportError:
    win32 = None


def carregar_planilha(caminho, sheet_name=None):
    """Carrega uma planilha Excel e retorna um DataFrame."""
    if not os.path.exists(caminho):
        print(f"Erro: O arquivo {caminho} não foi encontrado!")
        return None
    df = pd.read_excel(caminho, sheet_name=sheet_name)
    if isinstance(df, dict):  # Se retornar um dicionário, pega a primeira planilha
        df = list(df.values())[0]
    return df

def identificar_tipo_dado(entrada):
    """Identifica se o dado é CPF, Placa ou Ordem usando regex."""
    if re.fullmatch(r"\d{3}\.\d{3}\.\d{3}-\d{2}", entrada):
        return "CPF / CNPJ"
    elif re.fullmatch(r"[A-Z]{3}\d{4}", entrada) or re.fullmatch(r"[A-Z]{3}\d{1}[A-Z]{1}\d{2}", entrada):
        return "PLACA"
    elif entrada.isdigit():
        return "ORDEM"
    return None

def normalizar_placa(placa):
    """Normaliza a placa removendo hífen e espaços extras, e garantindo caixa alta."""
    if isinstance(placa, str):  
        return placa.replace("-", "").strip().upper()  # Remove hífen, espaços e deixa tudo maiúsculo
    return placa

def formatar_data(data):
    """Formata a data no formato DD/MM/YYYY"""
    return data.strftime("%d/%m/%Y") if hasattr(data, "strftime") else data

def formatar_hora(hora):
    """Formata a hora no formato HH:MM"""
    return hora.strftime("%H:%M") if hasattr(hora, "strftime") else hora

def obter_dados_diariovideo(placa):
    """Busca a DATA, HORA e DESCRIÇÃO na diariovideo.xlsx com base na PLACA, considerando variações de formatação."""
    if not os.path.exists(DIARIOVIDEO_FILE):
        print(f"Erro: O arquivo {DIARIOVIDEO_FILE} não foi encontrado!")
        return None

    df = carregar_planilha(DIARIOVIDEO_FILE, sheet_name="2025")
    if df is None:
        return None

    # Normalizar todas as placas da planilha
    df["PLACA"] = df["PLACA"].astype(str).apply(normalizar_placa)

    # Normalizar a placa de busca
    placa_normalizada = normalizar_placa(placa)

    # Filtrar os dados considerando a placa normalizada
    dados = df[df["PLACA"] == placa_normalizada]

    if dados.empty:
        print(f"Nenhuma infração encontrada para a placa {placa}.")
        return None
    
    linha = dados.iloc[0]
    local_raw = str(linha.get("LOCAL", "")).zfill(2)
    numero_camera = local_raw if local_raw in DADOS_CAMERA else None


    # Pegar a primeira ocorrência e formatar data/hora
    return {
        "DATA": formatar_data(linha["DATA"]),
        "HORA": formatar_hora(linha["HORA"]),
        "DESCRIÇÃO": linha["DESCRIÇÃO"],
        "CAMERA": numero_camera
    }

def obter_dados_pessoa(entrada):
    """Obtém os dados da pessoa na planilha plzdata.xlsx com base na entrada fornecida."""
    campo = identificar_tipo_dado(entrada)
    if not campo:
        print("Formato de dado inválido. Use CPF, Placa ou Número de Ordem.")
        return None
    
    df = carregar_planilha(PLZDATA_FILE)
    if df is None:
        return None
    
    df[campo] = df[campo].astype(str).str.strip()
    entrada = str(entrada).strip()
    
    dados_pessoa = df[df[campo] == entrada]
    return None if dados_pessoa.empty else dados_pessoa.iloc[0]

def preencher_celulas(ws, dados_pessoa, dados_video, numero_auto, fiscal):
    """Preenche a planilha com os dados da pessoa e da infração."""
    preencher_dados_pessoa(ws, dados_pessoa)
    preencher_dados_video(ws, dados_video)

    if dados_video and dados_video.get("CAMERA"):
        preencher_dados_camera(ws, dados_video["CAMERA"])

    preencher_dados_fiscal(ws, fiscal)

    if numero_auto:
        ws["F3"] = str(numero_auto).zfill(3)

def preencher_dados_camera(ws, numero_camera):
    """Preenche os dados de localização com base no número da câmera."""
    if numero_camera in DADOS_CAMERA:
        ws["D24"] = DADOS_CAMERA[numero_camera]["LOCALIZACAO"]
        ws["N26"] = DADOS_CAMERA[numero_camera]["BAIRRO"]
        ws["AL26"] = DADOS_CAMERA[numero_camera]["ZONA"]
        ws["E26"] = "S/N"

def preencher_dados_video(ws, dados_video):
    """Preenche os dados da infração se houver informações disponíveis."""
    if not dados_video:
        return
    ws["AE21"] = dados_video["DATA"]
    ws["AP21"] = dados_video["HORA"]
    ws["F40"] = dados_video["DESCRIÇÃO"].upper()
    ws["B40"] = "X"

def preencher_dados_pessoa(ws, dados_pessoa):
    """Preenche os dados do proprietário na planilha."""
    ws["B6"] = dados_pessoa["PROPRIETÁRIO"]
    ws["F42"] = f"FLAGRANTE REALIZADO PELO VIDEOMONITORAMENTO, VEÍCULO DE PLACA {dados_pessoa['PLACA']}"
    ws["AL15"] = dados_pessoa["CEP"]
    ws["D15"] = dados_pessoa["ENDEREÇO"]
    documento = dados_pessoa["CPF / CNPJ"]
    if is_cpf(documento):
        ws["B11"] = documento
        ws["AH11"] = None
    elif is_cnpj(documento):
        ws["AH11"] = documento
        ws["B11"] = None
    else:
        ws["B11"] = None
        ws["AH11"] = None
        print("Documento não reconhecido como CPF ou CNPJ.")

def is_cpf(documento):
    """Verifica se o documento está no formato de CPF."""
    return re.fullmatch(r"\d{3}\.\d{3}\.\d{3}-\d{2}", documento) is not None

def is_cnpj(documento):
    """Verifica se o documento está no formato de CNPJ."""
    return re.fullmatch(r"\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2}", documento) is not None


def preencher_dados_fiscal(ws, fiscal):
    """Preenche as células relacionadas ao fiscal."""
    codigo_fiscal, matricula_fiscal = buscar_fiscal(fiscal)
    if codigo_fiscal and matricula_fiscal:
        ws["B3"] = codigo_fiscal
        ws["B78"] = matricula_fiscal

def buscar_fiscal(fiscal):
    """Busca o código e matrícula do fiscal baseado no código ou matrícula."""
    if not fiscal:
        return None, None

    fiscal = fiscal.upper()  # Normaliza para caixa alta

    if fiscal in FISCAIS:
        return fiscal, FISCAIS[fiscal]["MATRICULA"]  # Retorna código e matrícula

    for codigo, info in FISCAIS.items():
        if info["MATRICULA"] == fiscal:
            return codigo, fiscal  # Retorna código correspondente e matrícula
    
    return None, None  # Caso não encontre

def preencher_auto(entrada, numero_auto=None, fiscal=None):
    """Cria uma cópia de auto.xlsx e preenche com os dados da pessoa e da infração."""
    dados_pessoa = obter_dados_pessoa(entrada)
    if dados_pessoa is None:
        print("Nenhuma pessoa encontrada com o dado fornecido.")
        return
    
    shutil.copy(AUTO_TEMPLATE_FILE, AUTO_OUTPUT_FILE)
    print(f"Cópia de {AUTO_TEMPLATE_FILE} criada como {AUTO_OUTPUT_FILE}")
    
    try:
        wb = load_workbook(AUTO_OUTPUT_FILE)
        ws = wb.active
    except Exception as e:
        print(f"Erro ao abrir a planilha copiada: {e}")
        return
    
    dados_video = obter_dados_diariovideo(dados_pessoa["PLACA"])
    preencher_celulas(ws, dados_pessoa, dados_video, numero_auto, fiscal)
    
    wb.save(AUTO_OUTPUT_FILE)
    print(f"Planilha preenchida: {AUTO_OUTPUT_FILE}")

def coletar_dados_terminal():
    """Coleta e valida os dados via terminal antes de preencher o auto."""
    print("=== Preenchimento de Auto de Infração ===")

    while True:
        entrada = input("Digite o CPF (000.000.000-00), CNPJ (00.000.000/0000-00), Placa (ABC1234) ou Nº de Ordem: ").strip().upper()
        if identificar_tipo_dado(entrada):
            break
        print("❌ Entrada inválida. Tente novamente com um formato reconhecido.")

    while True:
        numero_auto = input("Digite o número do auto (apenas números): ").strip()
        if numero_auto.isdigit():
            break
        print("❌ Número do auto inválido. Digite apenas números.")

    print("\n👮 Fiscais disponíveis:")
    for codigo, info in FISCAIS.items():
        print(f"  {codigo} (Matrícula: {info['MATRICULA']}): {info['NOME']}")

    while True:
        fiscal = input("\nDigite o código ou matrícula do fiscal: ").strip().upper()
        cod, mat = buscar_fiscal(fiscal)
        if cod and mat:
            fiscal = cod
            break
        print("❌ Código ou matrícula do fiscal não encontrados. Tente novamente.")

    preencher_auto(entrada, numero_auto, fiscal)

if __name__ == "__main__":
    #preencher_auto("2", "8", "05", "15133")  # Pode ser CPF, Placa ou Ordem
    coletar_dados_terminal()
